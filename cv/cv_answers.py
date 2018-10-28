import psycopg2
from openpyxl import Workbook


connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
cursor = connect.cursor()

cursor.execute("SELECT reports.metro_price,cv_responses.response,geo_objects.title \
FROM ma_metro.reports \
JOIN ma_metro.tasks \
	ON reports.task_id = tasks.id \
JOIN ma_metro.geo_objects \
	ON geo_objects.id = tasks.geo_object_id \
JOIN ma_metro.conveyor_jobs \
    ON reports.id = conveyor_jobs.target_id \
JOIN ma_metro.cv_responses \
    ON cv_responses.target_id = conveyor_jobs.id \
WHERE reports.state = 'accepted' \
    AND tasks.wave = 'os_auchan_42-18' \
    AND cv_responses.state = 'completed' \
    AND cv_responses.req_type = 'get_bulk_result_detect_classify_images'; \
    ")


data = list()

for i in cursor:
    data.append(i)

connect.close()

print(len(data))


book = Workbook()
sheet = book.active
sheet['A1'] = "report_price"
sheet['B1'] = "cv_price"
sheet['C1'] = "geo_object"
sheet['D1'] = "isequal"
sheet['E1'] = "url"



count = 0
for i in data:
    count += 1
    url = (i[1]['images'][0]['url'])
    # print(i[1]['images'][0]['tags'][0]['price_rub']['text'])
    if len(i[1]['images'][0]['tags']) == 0:
        # print('НЕТ ОТВЕТА ОТ СИВИ')

        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
        sheet.cell(row=count + 1, column=2).value = 'НЕТ ОТВЕТА ОТ СИВИ'
        sheet.cell(row=count + 1, column=4).value = 0
        sheet.cell(row=count + 1, column=1).value = i[0]
        sheet.cell(row=count + 1, column=3).value = i[2]
    elif len(i[1]['images'][0]['tags']) == 1:
        if i[1]['images'][0]['tags'][0]['price_rub']:
            if i[1]['images'][0]['tags'][0]['price_rub']['text'] != '':
                if i[1]['images'][0]['tags'][0]['price_kop']:
                    if i[1]['images'][0]['tags'][0]['price_kop']['text'] != '':
                        if i[0] == float(i[1]['images'][0]['tags'][0]['price_rub']['text'] + '.'\
                              + i[1]['images'][0]['tags'][0]['price_kop']['text']):
                            # print('СОВПАЛО!')
                            # записываем 1 в сравнение
                            sheet.cell(row=count + 1, column=2).value = str(i[1]['images'][0]['tags'][0]['price_rub']['text'] + '.'\
                              + i[1]['images'][0]['tags'][0]['price_kop']['text'])
                            sheet.cell(row=count + 1, column=4).value = 1
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        else:
                            # print(i[1]['images'][0]['tags'][0]['price_rub']['text'] + '.'\
                            # + i[1]['images'][0]['tags'][0]['price_kop']['text'])
                            sheet.cell(row=count + 1, column=2).value = str(
                                i[1]['images'][0]['tags'][0]['price_rub']['text'] + '.' \
                                + i[1]['images'][0]['tags'][0]['price_kop']['text'])
                            sheet.cell(row=count + 1, column=4).value = 0
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                    else:
                        if i[0] == int(i[1]['images'][0]['tags'][0]['price_rub']['text']):
                        # print(i[1]['images'][0]['tags'][0]['price_rub']['text'])
                            sheet.cell(row=count + 1, column=2).value = str(\
                                i[1]['images'][0]['tags'][0]['price_rub']['text'])
                            sheet.cell(row=count + 1, column=4).value = 1
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        else:
                            sheet.cell(row=count + 1, column=2).value = str(\
                            i[1]['images'][0]['tags'][0]['price_rub']['text'])
                            sheet.cell(row=count + 1, column=4).value = 0
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                else:
                    # print(i[1]['images'][0]['tags'][0]['price_rub']['text'])
                    if i[0] == int(i[1]['images'][0]['tags'][0]['price_rub']['text']):
                    # print(i[1]['images'][0]['tags'][0]['price_rub']['text'])
                        sheet.cell(row=count + 1, column=2).value = str(\
                                i[1]['images'][0]['tags'][0]['price_rub']['text'])
                        sheet.cell(row=count + 1, column=4).value = 1
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                    else:
                        sheet.cell(row=count + 1, column=2).value = str(\
                        i[1]['images'][0]['tags'][0]['price_rub']['text'])
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url

            else:
                if i[1]['images'][0]['tags'][0]['price_kop']:
                    if i[1]['images'][0]['tags'][0]['price_kop']['text'] != '':
                        # print('0' + '.' + i[1]['images'][0]['tags'][0]['price_kop']['text'])
                        if i[0] == float('0' + str(i[1]['images'][0]['tags'][0]['price_kop']['text'])):
                            sheet.cell(row=count + 1, column=4).value = 1
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=2).value = str(i[1]['images'][0]['tags'][0]['price_kop']['text'])
                        else:
                            sheet.cell(row=count + 1, column=4).value = 0
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                            sheet.cell(row=count + 1, column=2).value = str(
                            i[1]['images'][0]['tags'][0]['price_kop']['text'])


                    else:
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=2).value = 'НЕ СМОГ ОПРЕДЕЛИТЬ ЦЕНЫ'
                        # print('НЕТ ЦЕНЫ')
                else:
                    # print('НЕТ ЦЕНЫ')
                    sheet.cell(row=count + 1, column=4).value = 0
                    sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                    sheet.cell(row=count + 1, column=1).value = i[0]
                    sheet.cell(row=count + 1, column=3).value = i[2]
                    sheet.cell(row=count + 1, column=2).value = 'НЕ СМОГ ОПРЕДЕЛИТЬ ЦЕНЫ'
        else:
            if i[1]['images'][0]['tags'][0]['price_kop']:
                if i[1]['images'][0]['tags'][0]['price_kop']['text'] != '':
                    if i[0] == float('0' + str(i[1]['images'][0]['tags'][0]['price_kop']['text'])):
                        sheet.cell(row=count + 1, column=4).value = 1
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value = str(i[1]['images'][0]['tags'][0]['price_kop']['text'])
                    else:
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value = str(\
                        i[1]['images'][0]['tags'][0]['price_kop']['text'])
                    # print('0' + '.' + i[1]['images'][0]['tags'][0]['price_kop']['text'])
                else:
                    sheet.cell(row=count + 1, column=4).value = 0
                    sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                    sheet.cell(row=count + 1, column=1).value = i[0]
                    sheet.cell(row=count + 1, column=3).value = i[2]
                    sheet.cell(row=count + 1, column=2).value = 'НЕ СМОГ ОПРЕДЕЛИТЬ ЦЕНЫ'
                    # print('НЕТ ЦЕНЫ')
            else:
                sheet.cell(row=count + 1, column=4).value = 0
                sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                sheet.cell(row=count + 1, column=1).value = i[0]
                sheet.cell(row=count + 1, column=3).value = i[2]
                sheet.cell(row=count + 1, column=2).value = 'НЕ СМОГ ОПРЕДЕЛИТЬ ЦЕНЫ'

    else:
        # print('2 ЦЕНЫ')
        for j in range(len(i[1]['images'][0]['tags'])-1):
            max_prob = 0
            mx_prob_index = 0
            if i[1]['images'][0]['tags'][j]['prob'] > max_prob:
                max_prob = i[1]['images'][0]['tags'][j]['prob']
                max_prob_index = j
            #     print(j,'INDEEEEX')
            # print(j, 'INDEEEEX ^^')
        if i[1]['images'][0]['tags'][max_prob_index]['price_rub']:
            if i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'] != '':
                if i[1]['images'][0]['tags'][max_prob_index]['price_kop']:
                    if i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'] != '':
                        if i[0] == float(i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'] + '.'\
                              + i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text']):
                            # print('СОВПАЛО!')
                            # записываем 1 в сравнение
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                            sheet.cell(row=count + 1, column=4).value = 1
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=2).value = str(i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'] + '.'\
                              + i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                            sheet.cell(row=count + 1, column=3).value = i[2]
                        else:
                            # print(i[1]['images'][0]['tags'][j]['price_rub']['text'] + '.'\
                            # + i[1]['images'][0]['tags'][j]['price_kop']['text'])
                            sheet.cell(row=count + 1, column=4).value = 0
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=2).value = str(i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'] + '.'\
                            + i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                            sheet.cell(row=count + 1, column=3).value = i[2]
                    else:
                        # print(i[1]['images'][0]['tags'][j]['price_rub']['text'])
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value =i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text']
                else:
                    # print(i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'])
                    if i[0] == int(str(i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'])):
                        sheet.cell(row=count + 1, column=4).value = 1
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value = str(i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'])
                    else:
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value = str(\
                        i[1]['images'][0]['tags'][max_prob_index]['price_rub']['text'])

            else:
                if i[1]['images'][0]['tags'][max_prob_index]['price_kop']:
                    if i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'] != '':
                        # print('0' + '.' + i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                        if i[0] == float('0' + '.' + str(i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])):
                            sheet.cell(row=count + 1, column=4).value = 1
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=2).value = str(\
                                i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                        else:
                            sheet.cell(row=count + 1, column=4).value = 0
                            sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                            sheet.cell(row=count + 1, column=1).value = i[0]
                            sheet.cell(row=count + 1, column=3).value = i[2]
                            sheet.cell(row=count + 1, column=2).value = str( \
                                i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                    else:
                        # print('НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ')
                        sheet.cell(row=count + 1, column=2).value = 'НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ'
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                else:
                    # print('НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ')
                    sheet.cell(row=count + 1, column=2).value = 'НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ'
                    sheet.cell(row=count + 1, column=4).value = 0
                    sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                    sheet.cell(row=count + 1, column=1).value = i[0]
                    sheet.cell(row=count + 1, column=3).value = i[2]
        else:
            if i[1]['images'][0]['tags'][max_prob_index]['price_kop']:
                if i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'] != '':
                    # print('0' + '.' + i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                    if i[0] == float('0' + '.' + str(i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])):
                        sheet.cell(row=count + 1, column=4).value = 1
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value = str( \
                            i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])
                    else:
                        sheet.cell(row=count + 1, column=4).value = 0
                        sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                        sheet.cell(row=count + 1, column=1).value = i[0]
                        sheet.cell(row=count + 1, column=3).value = i[2]
                        sheet.cell(row=count + 1, column=2).value = str(\
                            i[1]['images'][0]['tags'][max_prob_index]['price_kop']['text'])

                else:
                    # print('НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ')
                    sheet.cell(row=count + 1, column=2).value = 'НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ'
                    sheet.cell(row=count + 1, column=4).value = 0
                    sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                    sheet.cell(row=count + 1, column=1).value = i[0]
                    sheet.cell(row=count + 1, column=3).value = i[2]
            else:
                # print('НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ')
                sheet.cell(row=count + 1, column=2).value = 'НЕТ ЦЕНЫ ПОД ЭТИМ ИНДЕКСОМ'
                sheet.cell(row=count + 1, column=4).value = 0
                sheet.cell(row=count + 1, column=5).value = '=HYPERLINK("%s")' % url
                sheet.cell(row=count + 1, column=1).value = i[0]
                sheet.cell(row=count + 1, column=3).value = i[2]



connect.close()
book.save('xxx.xlsx')

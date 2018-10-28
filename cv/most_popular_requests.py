import pandas as pd
from openpyxl import Workbook


df = pd.read_csv('data_prod.csv',sep='|')


days = df['date'].value_counts()
mp = df['url'].value_counts().head(20)


book = Workbook()
sheet = book.active

sheet['A1'] = "DATE"
sheet['B1'] = "URL"
sheet['C1'] = "MIN"
sheet['D1'] = "MAX"
sheet['E1'] = "MEAN"
sheet['F1'] = "TOTAL"




c = 0
for d in sorted(days.keys()):
    for k in mp.keys():
        c += 1
        print(d,k,sep='===>')
        sheet.cell(row=c+1,column=1).value = d
        sheet.cell(row=c+1,column=2).value = k
        sheet.cell(row=c+1,column=3).value = df[(df['url'] == '{}'.format(k)) & (df['date'] == '{}'.format(d))]['spent'].min()
        sheet.cell(row=c+1,column=4).value = df[(df['url'] == '{}'.format(k)) & (df['date'] == '{}'.format(d))]['spent'].max()
        sheet.cell(row=c+1,column=5).value = df[(df['url'] == '{}'.format(k)) & (df['date'] == '{}'.format(d))]['spent'].mean()
        sheet.cell(row=c+1,column=6).value = df[(df['url'] == '{}'.format(k)) & (df['date'] == '{}'.format(d))]['spent'].count()


book.save('top_requests_for_days_prod.xlsx')

import psycopg2
from openpyxl import Workbook

connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
cursor = connect.cursor()

cursor.execute("SELECT rr.metro_price,cv_responses.response,geo_objects.title,\
    rr.barcode,rr.sku_title, cv_responses.response->'images'->0->'tags'->0->'titles' \
    FROM ma_metro.reports r \
        JOIN ma_metro.tasks \
    ON r.task_id = tasks.id \
        JOIN ma_metro.geo_objects \
    ON geo_objects.id = tasks.geo_object_id \
        JOIN ma_metro.conveyor_jobs \
    ON r.id = conveyor_jobs.target_id \
        JOIN ma_metro.cv_responses \
    ON cv_responses.target_id = conveyor_jobs.id \
        JOIN ma_metro.reports rr \
    ON r.id = rr.origin_report_id \
    WHERE cv_responses.state = 'completed' \
        AND rr.state = 'accepted' \
        AND conveyor_jobs.wave = 'magnit-cv-3' \
        AND cv_responses.req_type = 'get_bulk_result_detect_classify_images';")



arr = list()
book = Workbook()
sheet = book.active

sheet['A1'] = "report_price"
sheet['B1'] = "cv_response.body"
sheet['C1'] = "geo_object"
sheet['D1'] = "barcode"
sheet['E1'] = "sku_title"
sheet['F1'] = "tags-titles"
sheet['G1'] = "url"



for i in cursor:
    arr.append(i)

count = 0
for i in arr:
    count += 1
    if len(i[5]) == 0:
        sheet.cell(row=count+1, column=1).value = i[0]
        sheet.cell(row=count+1, column=2).value = str(i[1])
        sheet.cell(row=count+1, column=3).value = i[2]
        sheet.cell(row=count+1, column=4).value = i[3]
        sheet.cell(row=count+1, column=5).value = i[4]
        sheet.cell(row=count+1, column=6).value = str(i[5])
        sheet.cell(row=count+1,column=7).value = i[1]['images'][0]['url']
    elif len(i[5]) == 1:
        sheet.cell(row=count+1, column=1).value = i[0]
        sheet.cell(row=count+1, column=2).value = str(i[1])
        sheet.cell(row=count+1, column=3).value = i[2]
        sheet.cell(row=count+1, column=4).value = i[3]
        sheet.cell(row=count+1, column=5).value = i[4]
        sheet.cell(row=count+1, column=6).value = i[5][0]['text']
        sheet.cell(row=count+1,column=7).value = i[1]['images'][0]['url']
    else:
        title = str()
        sorted_titles = sorted(i[5],key=lambda x: x['vertexes'][0]['y'])
        for x in sorted_titles:
            title = (title + ' ' + x['text']).strip()

        sheet.cell(row=count+1, column=1).value = i[0]
        sheet.cell(row=count+1, column=2).value = str(i[1])
        sheet.cell(row=count+1, column=3).value = i[2]
        sheet.cell(row=count+1, column=4).value = i[3]
        sheet.cell(row=count+1, column=5).value = i[4]
        sheet.cell(row=count+1, column=6).value = title
        sheet.cell(row=count+1,column=7).value = i[1]['images'][0]['url']

book.save('titles-concat.xlsx')
connect.close()

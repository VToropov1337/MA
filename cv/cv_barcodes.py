import psycopg2
from openpyxl import Workbook

connect = psycopg2.connect(database=***, user=***, host=***, port=***, password=***)
cursor = connect.cursor()


cursor.execute("SELECT r.id, r.metro_price,r.barcode,geo_objects.title, cv_responses.response->'images'->0->'tags'->0->'barcode',cv_responses.response->'images'->0->'tags'->0->'barcode_num'\
    FROM ma_metro.reports r \
    JOIN ma_metro.reports rr \
        ON rr.id = r.origin_report_id \
        JOIN ma_metro.tasks \
    ON rr.task_id = tasks.id \
        JOIN ma_metro.geo_objects \
    ON geo_objects.id = tasks.geo_object_id \
        JOIN ma_metro.conveyor_jobs \
    ON r.id = conveyor_jobs.target_id \
        JOIN ma_metro.cv_responses \
    ON cv_responses.target_id = conveyor_jobs.id \
    WHERE cv_responses.state = 'completed' \
        AND r.state = 'accepted'\
        AND conveyor_jobs.wave = 'os_auchan_44-18' \
        AND cv_responses.state = 'completed' \
        AND cv_responses.req_type = 'get_bulk_result_detect_classify_images';")


arr = list()


for i in cursor:
    arr.append(i)

# print(len(arr))




book = Workbook()
sheet = book.active

sheet['A1'] = "report_id"
sheet['B1'] = "report_price"
sheet['C1'] = "report_barcode"
sheet['D1'] = "geo_object"
sheet['E1'] = "cv_barcode"
sheet['F1'] = "cv_barcode_num"
sheet['G1'] = "isequal"





c = 0
for j in arr:
    c += 1
    print(c)
    if j[4] == None:
        sheet.cell(row=c+1, column=5).value = 'Не прочитан'
        sheet.cell(row=c+1, column=7).value = 0
    else:
        sheet.cell(row=c+1, column=5).value = j[4]['text']
        if j[2] == j[4]['text']:
            sheet.cell(row=c+1, column=7).value = 1
        else:
            sheet.cell(row=c+1, column=7).value = 0


    if j[5] == None:
        sheet.cell(row=c+1, column=6).value = 'Не прочитан'
    else:
        sheet.cell(row=c+1, column=6).value = j[5]['text']


    sheet.cell(row=c+1, column=1).value = j[0]
    sheet.cell(row=c+1, column=2).value = j[1]
    sheet.cell(row=c+1, column=3).value = j[2]
    sheet.cell(row=c+1, column=4).value = j[3]


book.save('44-18-barcode.xlsx')
connect.close()

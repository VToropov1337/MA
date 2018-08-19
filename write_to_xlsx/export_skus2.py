import psycopg2
from openpyxl import Workbook

arr = ['equal','equal_with_analog','analog_with_recalculation','analog_without_recalculation']

connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')


for i in arr:
    counter = 1
    cursor = connect.cursor()
    cursor.execute("SELECT id,metro_title, metro_description,subsys_art_no, metro_collection_method \
    FROM ma_metro.articles \
    WHERE company_id = 8 \
    AND metro_collection_method = %(name)s LIMIT 9;", {'name': i})

    book = Workbook()
    sheet = book.active
    base = list()

    for elem in cursor:
        print(elem)
        print('---')
        sheet.cell(row=counter, column=1).value = elem[0]
        sheet.cell(row=counter, column=2).value = elem[1]
        sheet.cell(row=counter, column=3).value = elem[2]
        sheet.cell(row=counter, column=4).value = elem[3]
        sheet.cell(row=counter, column=5).value = elem[4]
        counter +=1
    book.save(('write_skus_{}.xlsx').format(i))

connect.close()

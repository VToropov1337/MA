import psycopg2
from openpyxl import Workbook

arr = ['equal','equal_with_analog','analog_with_recalculation','analog_without_recalculation']

connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
counter = 1


for i in arr:
    cursor = connect.cursor()
    cursor.execute("SELECT id,metro_title, metro_description,subsys_art_no, metro_collection_method \
    FROM ma_metro.articles \
    WHERE company_id = 8 \
    AND metro_collection_method = %(name)s LIMIT 9;", {'name': i})

    book = Workbook()
    sheet = book.active
    base = list()

    for elem in cursor:
        base.append(elem)


    for j in range(len(base)):

        sheet.cell(row=1+j, column=1).value = base[j][0]
        sheet.cell(row=1+j, column=2).value = base[j][1]
        sheet.cell(row=1+j, column=3).value = base[j][2]
        sheet.cell(row=1+j, column=4).value = base[j][3]
        sheet.cell(row=1+j, column=5).value = base[j][4]
        counter += 1 

    book.save(('write_skus_{}.xlsx').format(i))

connect.close()

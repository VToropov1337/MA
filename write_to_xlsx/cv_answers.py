import psycopg2
from openpyxl import Workbook
import csv

connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
cursor = connect.cursor()
cursor.execute("SELECT reports.metro_price,cv_responses.response,geo_objects.title \
FROM ma_metro.reports \
JOIN ma_metro.tasks \
	ON reports.task_id = tasks.id \
JOIN ma_metro.geo_objects \
	ON geo_objects.id = tasks.geo_object_id \
JOIN ma_metro.conveyor_jobs \
    ON reports.id = conveyor_jobs.target_id \
JOIN ma_metro.cv_responses \
    ON cv_responses.target_id = conveyor_jobs.id \
WHERE reports.state = 'accepted' \
    AND conveyor_jobs.wave = '***' \
    AND cv_responses.state = 'completed' \
    AND cv_responses.req_type = 'get_bulk_result_detect_classify_images';")

arr = list()

for i in cursor:
    arr.append(i)


book = Workbook()
sheet = book.active


print(arr[0][1]['images'][0]['tags'][0]['price_rub']['text'])
count = 0
for i in arr:
    count += 1
    if i[1]['images'][0]['tags'] != []: #если есть таги, идем дальше
        if i[1]['images'][0]['tags'][0]['price_rub'] != None: #eсли цена не нан
            if i[1]['images'][0]['tags'][0]['price_kop'] != None:
                print('хорошо, есть полная цена')
                z = i[0]
                sheet.cell(row=count, column=1).value = z
                k = str(i[1]['images'][0]['tags'][0]['price_rub']['text']+','+i[1]['images'][0]['tags'][0]['price_kop']['text'])
                sheet.cell(row=count, column=2).value = k
                sheet.cell(row=count, column=3).value = i[2]

            else:
                print('неплохо, есть только цена')
                z = i[0]
                sheet.cell(row=count, column=1).value = z
                k = str(i[1]['images'][0]['tags'][0]['price_rub']['text'])
                sheet.cell(row=count, column=2).value = k
                sheet.cell(row=count, column=3).value = i[2]

        else:
            print('цена отсутствует')
            z = i[0]
            sheet.cell(row=count, column=1).value = z
            sheet.cell(row=count, column=3).value = i[2]
    else:
        print('плохо')
        z = i[0]
        sheet.cell(row=count, column=1).value = z
        sheet.cell(row=count, column=3).value = i[2]


book.save('prices.xlsx')
connect.close()

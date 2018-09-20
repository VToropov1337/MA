import psycopg2
from openpyxl import Workbook

connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
cursor = connect.cursor()

cursor.execute("SELECT users.id, users.phone FROM m10.users \
JOIN m10.groups_members \
    ON users.id = groups_members.member_id \
JOIN m10.groups \
    ON groups.id = groups_members.group_id \
WHERE group_id = 6;")

data = list()
for i in cursor:
    data.append(i)


cursor.execute("SELECT DISTINCT(users.id), users.phone FROM m10.users \
JOIN m10.jobs_decisions\
    ON users.id = jobs_decisions.user_id;")


data2 = list()
for i in cursor:
    data2.append(i)

for i in data2:
    if i not in data:
        data.append(i)




book = Workbook()
sheet = book.active
sheet['A1'] = "user_id"
sheet['B1'] = "phone"

count = 0


for i in data:
    count += 1
    sheet.cell(row=count+1,column = 1).value = i[0]
    sheet.cell(row=count+1,column = 2).value = i[1]




book.save('m10.xlsx')
connect.close()


connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
cursor = connect.cursor()

cursor.execute("SELECT users.id, users.phone FROM ma_metro.users \
JOIN ma_metro.groups_members \
    ON users.id = groups_members.member_id \
JOIN ma_metro.groups \
    ON groups.id = groups_members.group_id \
WHERE group_id = 33;")

data = list()
for i in cursor:
    data.append(i)


cursor.execute("SELECT DISTINCT(users.id), users.phone FROM ma_metro.users \
JOIN ma_metro.conveyor_jobs_decisions\
    ON users.id = conveyor_jobs_decisions.user_id;")


data2 = list()
for i in cursor:
    data2.append(i)

for i in data2:
    if i not in data:
        data.append(i)




book = Workbook()
sheet = book.active
sheet['A1'] = "user_id"
sheet['B1'] = "phone"

count = 0


for i in data:
    count += 1
    sheet.cell(row=count+1,column = 1).value = i[0]
    sheet.cell(row=count+1,column = 2).value = i[1]




book.save('m54.xlsx')
connect.close()

import psycopg2
from openpyxl import Workbook


# Выгрузка из старой поддержки по обращениям проекта
# Иванов также работает с текущей поддержкой по проекту. Нужно понять объем работ и время когда
# обращения сыпятся. Инфа нужна за весь июль 18. Формат:
# Дата и время создания заявки / Дата и время завершения заявки / Текст заявки



connect = psycopg2.connect(database='***', user='psqlreader', host='***', port='***', password='***')
cursor = connect.cursor()

cursor.execute(
    "SELECT * FROM tickets t \
        LEFT JOIN (SELECT DISTINCT ON(ticket_id)* FROM ticket_comments ORDER BY ticket_id, id ASC) AS commentz \
        ON t.id = commentz.ticket_id \
        WHERE t.state = 'closed' \
        AND t.created_at >= '2018/07/01' \
        AND t.created_at <= '2018/08/01';"
)


book = Workbook()
sheet = book.active


base = list()
for elem in cursor:
    print(elem)
    base.append(elem)


for j in range(len(base)):
    sheet.cell(row=j+1, column=1).value = base[j][5]
    sheet.cell(row=j+1, column=2).value = base[j][6]
    sheet.cell(row=j+1, column=3).value = base[j][9]
    book.save('write_projectX.xlsx')

connect.close()

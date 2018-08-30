import psycopg2
from openpyxl import Workbook

connect = psycopg2.connect(database='***', user='psqlreader', host='***', port='***', password='***')
cursor = connect.cursor()


cursor.execute("SELECT DISTINCT moderator_id, CONCAT_WS(' ',first_name, last_name) as fullname, COUNT(*) as qty, moderated_at  FROM ma_metro.reports r \
    JOIN ma_metro.tasks t \
    ON r.task_id = t.id \
    JOIN ma_metro.users u \
    ON r.moderator_id = u.id \
    WHERE t.wave = 'm54_2518_btf' AND r.metro_available = False \
    AND r.state = 'accepted' \
    GROUP BY moderator_id, first_name, last_name, moderated_at;")


arr = list()
book = Workbook()
sheet = book.active

for i in cursor:
    arr.append(i)

count = 0
for j in range(len(arr)):
    sheet.cell(row=j+1, column=1).value = arr[j][0]
    sheet.cell(row=j+1, column=2).value = arr[j][1]
    sheet.cell(row=j+1, column=3).value = arr[j][2]
    sheet.cell(row=j+1, column=4).value = arr[j][3]
    count += 1
    print(count)

book.save('moderators.xlsx')
connect.close()

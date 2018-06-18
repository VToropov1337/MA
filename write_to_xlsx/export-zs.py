import psycopg2
from openpyxl import Workbook


connect = psycopg2.connect(database='postgres', user='master', host='localhost', password='master')
cursor = connect.cursor()
cursor.execute("SELECT article_title, articles.price, geo_objects.title \
FROM articles \
JOIN geo_objects \
ON articles.geo_object_id = geo_objects.id;")


book = Workbook()
sheet = book.active


sheet['A1'] = "Sku"

base = list()
uniq_sku = list()
uniq_tt = list()

for elem in cursor:
    base.append([elem[0],elem[1],elem[2]])

# base = [
# ['Молоко Домик в деревне 3,2% 1л', 42.5, 'Verniy']
# ['Молоко Домик в деревне 3,2% 1л', 42.5, 'Okey']
# ['Молоко Домик в деревне 3,2% 1л', 42.4, 'Metro']
# ['Молоко Parmalat 3,2% 1л', 42.5, 'Perekrestok']
# ['Молоко Parmalat 3,2% 1л', 42.5, 'Dixy']
# ['Шоколадный батончик Snikers 80 г', 18.5, 'Selgros']
# ['Шоколадный батончик Snikers 80 г', 18.5, 'Auchan']
# ['Сыр Пешехонский 1кг', 380.0, 'Okey']
# ['Сыр Пешехонский 1кг', 350.0, 'Metro']
# ['Сыр Пешехонский 1кг', 390.0, 'Auchan']
# ]


#без счетчика, каждый раз руками прибавляю + 1

for i in base:
    if i[0] not in uniq_sku:
        uniq_sku.append(i[0])
        sheet.cell(row=1+uniq_sku.index(i[0])+1, column=1).value = i[0]

    if i[2] not in uniq_tt:
        uniq_tt.append(i[2])
        sheet.cell(row=1, column=uniq_tt.index(i[2])+1+1).value = i[2]

    sheet.cell(row=uniq_sku.index(i[0])+1+1, column=uniq_tt.index(i[2])+1+1).value = i[1] #прибавляю дополнительную единицу, тк массив индексируется с 0

book.save('write2cell.xlsx')
connect.close()

import psycopg2
from openpyxl import Workbook

connect = psycopg2.connect(database='postgres', user='master', host='localhost', password='master')
cursor = connect.cursor()

cursor.execute("SELECT article_title, articles.price, geo_objects.title \
FROM articles \
JOIN geo_objects \
ON articles.geo_object_id = geo_objects.id \
GROUP BY geo_objects.title,article_title,articles.price;")

book = Workbook()
sheet = book.active
sheet['A1'] = "Sku"

#создаю счетчик
i = 1

#иду построчно по объекту и заполняю каждую строчку
for line in cursor:
    sheet.cell(row=1+i, column=1).value = line[0]
    sheet.cell(row=1+i, column=i+1).value = line[1]
    sheet.cell(row=1, column=i+1).value = line[2]
    i += 1

book.save('write2cell.xlsx')

connect.close()

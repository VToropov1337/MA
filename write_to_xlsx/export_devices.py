import psycopg2
from openpyxl import Workbook
import pandas as pd


connect = psycopg2.connect(database='***', user='***', host='***', port='***', password='***')
cursor = connect.cursor()


cursor.execute("SELECT  u.id as user_id, CONCAT_WS(' ',u.first_name,u.last_name) as fullname,u.roles, d.platform,d.specs->'device',d.specs->'os' \
    FROM ma_metro.users u \
JOIN ma_metro.reports r \
	ON r.user_id = u.id \
JOIN ma_metro.devices d \
	ON d.user_id = u.id \
JOIN ma_metro.tasks \
	ON tasks.id = r.task_id \
WHERE tasks.project_id = 1 \
AND r.created_at::text LIKE '2018%' \
AND platform != 'browser' \
GROUP BY (u.id,d.platform,d.specs->'device',d.specs->'os') \
")


data = list()

for i in cursor:
    data.append(i)


book = Workbook()
sheet = book.active
sheet['A1'] = "user_id"
sheet['B1'] = "fullname"
sheet['C1'] = "roles"
sheet['D1'] = "os"
sheet['E1'] = "model"
sheet['F1'] = "version"


count = 0
for j in data:
    count += 1
    print(count)
    sheet.cell(row=count+1, column=1).value = j[0]
    sheet.cell(row=count+1, column=2).value = j[1]
    sheet.cell(row=count+1, column=3).value = str(j[2])
    sheet.cell(row=count+1, column=4).value = j[3]
    sheet.cell(row=count+1, column=5).value = j[4]
    sheet.cell(row=count+1, column=6).value = j[5]


book.save('agents.xlsx')
connect.close()


file = 'agents.xlsx'
xl = pd.ExcelFile(file)
df1 = xl.parse('Sheet')
df1.to_csv('agents.csv')

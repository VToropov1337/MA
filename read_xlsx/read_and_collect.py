from openpyxl import load_workbook

wb = load_workbook('./users_coordinate.xlsx')
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')

data = list()
for i in range(2,10):
     hh = dict()
     hh['user_id'] = sheet.cell(row=i, column=1).value
     hh['phone_number'] = sheet.cell(row=i, column=2).value
     hh['ltd'] = sheet.cell(row=i, column=3).value
     hh['lgt'] = sheet.cell(row=i, column=4).value
     data.append(hh)

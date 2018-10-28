from openpyxl import load_workbook
import psycopg2


connect = psycopg2.connect(database='postgres', user='master', host='localhost', password='master')
cursor = connect.cursor()

wb = load_workbook('./users_coordinate.xlsx')
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')

data = list()
for i in range(2,100):
    if sheet.cell(row=i, column=1).value != None:
        hh = dict()
        hh['user_id'] = sheet.cell(row=i, column=1).value
        hh['phone_number'] = sheet.cell(row=i, column=2).value
        hh['ltd'] = sheet.cell(row=i, column=3).value
        hh['lgt'] = sheet.cell(row=i, column=4).value
        data.append(hh)





for i in data:
    cursor.execute(
        """INSERT INTO public.users (id,phone,ltd,lgt) VALUES (%(id)s,%(phone)s,%(ltd)s,%(lgt)s)""", \
                    {'id': i['user_id'], 'phone': i['phone_number'], 'ltd': i['ltd'],'lgt': i['lgt']} \
                    )
    connect.commit()
    print('added {}'.format(i))

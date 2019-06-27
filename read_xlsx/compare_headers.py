from openpyxl import load_workbook
import os

def create_list_of_files():
    all_files = list()
    rootdir = os.getcwd()

    for root, dirs, files in os.walk(rootdir):
        for file in files:
            if file.endswith(".xlsx"):
                all_files.append(os.path.join(root, file))

    return all_files


def create_ethalon(path_to_file):
    wb = load_workbook(path_to_file)
    print(wb.sheetnames)
    sheet = wb['здание_полное']

    ethalon = []
    for cell in list(sheet.rows)[1]:
        if cell.value == None:
            continue
        ethalon.append(cell.value)
    return ethalon


def compare_headers(file, ethalon):
    wb = load_workbook(file)
    sheet = wb['здание_полное']

    headers = []
    for cell in list(sheet.rows)[1]:
        if cell.value == None:
            continue
        headers.append(cell.value)

    diff = []
    for i in headers:
        if i not in ethalon:
            diff.append(i)

    if len(diff) == 0:
        print('Файл: ' + file + ' совпадает с эталоном')
    else:
        print('Файл: ' + file,'Неожиданные колонки: ',diff)


    
def check_files(files,ethalon):
    for i in files:
        compare_headers(i, ethalon)


if __name__ == '__main__':
    files = create_list_of_files()
    check_files(files,create_ethalon('path_to_ethalon_file')) 